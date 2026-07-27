from pathlib import Path
import re
import csv
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook

# ----------------------------
# Seleções
# ----------------------------
def selecionar_entrada_saida_e_base():
    root = tk.Tk()
    root.withdraw()

    entrada = filedialog.askdirectory(title="Pasta de entrada")
    if not entrada:
        return None, None, None

    saida = filedialog.askdirectory(title="Pasta de saída")
    if not saida:
        return None, None, None

    base = filedialog.askopenfilename(
        title="Base de nomenclatura",
        filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
    )
    if not base:
        return None, None, None

    return Path(entrada), Path(saida), Path(base)

# ----------------------------
# Popup simples
# ----------------------------
def perguntar_ordem_popup(letras):
    root = tk.Tk()
    root.withdraw()

    letras_str = ", ".join(sorted(letras))

    ordem = simpledialog.askstring(
        "Ordem dos PDFs",
        f"Letras encontradas: {letras_str}\n\nDigite a ordem (ex: C,G,A):"
    )

    if not ordem:
        return None

    return [x.strip().upper() for x in ordem.split(",") if x.strip()]

# ----------------------------
# Utilidades
# ----------------------------
def normalizar_nome(nome):
    return re.sub(r"\.pdf$", "", str(nome).strip(), flags=re.IGNORECASE)

def sanitizar(nome):
    nome = re.sub(r'[<>:"/\\|?*]', "_", str(nome))
    return re.sub(r"\s+", " ", nome).strip()

def extrair_letra(nome):
    m = re.match(r"^\d+([a-z])", nome.lower())
    return m.group(1).upper() if m else ""

# ----------------------------
# Base
# ----------------------------
def carregar_base(caminho):
    mapa = {}

    if caminho.suffix.lower() == ".csv":
        with open(caminho, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for linha in reader:
                if len(linha) >= 2:
                    k = normalizar_nome(linha[0]).lower()
                    v = sanitizar(linha[1])
                    if k and v:
                        mapa[k] = v
    else:
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 2:
                k = normalizar_nome(row[0]).lower()
                v = sanitizar(row[1])
                if k and v:
                    mapa[k] = v

    return mapa

# ----------------------------
# Ordem dinâmica
# ----------------------------
def definir_ordem(nome, ordem_usuario):
    nome = nome.lower()

    m = re.match(r"^(\d+)([a-z]?)(\d*)\.pdf$", nome)
    if not m:
        return 9999

    letra = m.group(2).upper()
    numero = m.group(3)

    if ordem_usuario and letra in ordem_usuario:
        pos = ordem_usuario.index(letra)
    else:
        pos = 999

    num_extra = int(numero) if numero else 0

    return pos * 100 + num_extra

# ----------------------------
# Nome final
# ----------------------------
def obter_nome_saida(codigo, arquivos, mapa):
    candidatos = [codigo]

    for arq in arquivos:
        candidatos.append(normalizar_nome(arq.name))

    for c in candidatos:
        if c.lower() in mapa:
            return sanitizar(mapa[c.lower()])

    return codigo

# ----------------------------
# Evitar sobrescrever
# ----------------------------
def nome_unico(path):
    if not path.exists():
        return path

    i = 1
    while True:
        novo = path.with_name(f"{path.stem}_{i}{path.suffix}")
        if not novo.exists():
            return novo
        i += 1

# ----------------------------
# MAIN
# ----------------------------
def main():
    pasta_entrada, pasta_saida, base = selecionar_entrada_saida_e_base()
    if not pasta_entrada:
        return

    pasta_saida.mkdir(exist_ok=True)

    try:
        mapa = carregar_base(base)
    except Exception as e:
        messagebox.showerror("Erro", str(e))
        return

    padrao = re.compile(r"^(\d+)([a-z]?)(\d*)\.pdf$", re.IGNORECASE)

    grupos = {}
    letras_detectadas = set()
    ignorados = []

    # leitura dos arquivos
    for arq in pasta_entrada.iterdir():
        if not (arq.is_file() and arq.suffix.lower() == ".pdf"):
            continue

        nome = arq.name.strip()
        m = padrao.match(nome)

        if not m:
            ignorados.append(nome)
            continue

        codigo = m.group(1)

        letra = extrair_letra(nome)
        if letra:
            letras_detectadas.add(letra)

        grupos.setdefault(codigo, []).append(arq)

    if not grupos:
        messagebox.showinfo("Resultado", "Nenhum PDF válido encontrado.")
        return

    # pergunta ordem
    ordem_usuario = []
    if letras_detectadas:
        ordem_usuario = perguntar_ordem_popup(letras_detectadas)

        if not ordem_usuario:
            messagebox.showwarning("Aviso", "Ordem não definida.")
            return

    gerados = []

    # merge
    for codigo, arquivos in grupos.items():
        try:
            arquivos.sort(key=lambda x: definir_ordem(x.name, ordem_usuario))

            writer = PdfWriter()

            for arq in arquivos:
                reader = PdfReader(str(arq))
                for pagina in reader.pages:
                    writer.add_page(pagina)

            nome_saida = obter_nome_saida(codigo, arquivos, mapa)
            destino = nome_unico(pasta_saida / f"{nome_saida}.pdf")

            with open(destino, "wb") as f:
                writer.write(f)

            gerados.append(destino.name)

        except Exception as e:
            print(f"Erro em {codigo}: {e}")

    messagebox.showinfo("Concluído", f"{len(gerados)} PDFs gerados")

    print("\nGerados:")
    for g in gerados:
        print(g)

    if ignorados:
        print("\nIgnorados:")
        for i in ignorados:
            print(i)

# ----------------------------
if __name__ == "__main__":
    main()