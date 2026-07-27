"""
Unificador de PDFs / Notas de Débito - v5.0
=============================================

Fluxo:
 1. Lê os comprovantes PDF de uma pasta de entrada (padrão de nome:
    "<código><letra opcional><número opcional>.pdf", ex: "100.pdf", "100a.pdf").
 2. Lê também os arquivos de NOTA DE DÉBITO: qualquer PDF cujo nome comece
    com "Nota de débito" (independente da marca/texto que vem depois, ex:
    "Nota de débito - Neon Consiga+ ND 1266.pdf",
    "Nota de débito - Neon Pagamentos ND 1262.pdf"). O número do ND e o
    "prefixo" (tudo antes de "ND") são extraídos do próprio nome do
    arquivo.
 3. Lê uma base (Excel/CSV). São aceitos dois formatos:

    a) Formato NOVO (com cabeçalho), colunas (em qualquer ordem):
         "Id"                 -> código do arquivo (ex: 100)
         "LANÇAMENTO"         -> nome/descrição (usado só no fallback sem ND)
         "DATA DO PAGAMENTO"  -> data usada para ordenar os códigos dentro
                                  de um mesmo ND (mais antiga primeiro)
         "ND"                 -> número da nota de débito

    b) Formato ANTIGO (sem cabeçalho, compatibilidade com versões
       anteriores): Coluna A = código, Coluna B = nome, Coluna C = ND
       (sem data de pagamento).

 4. Para cada número de ND presente na base, gera UM ÚNICO PDF unificado
    contendo: o arquivo da nota de débito daquele ND (sempre como
    PRIMEIRA página) + todos os comprovantes de todos os códigos
    vinculados àquele ND, ordenados por:
        1º) posição da letra do arquivo na ordem informada pelo usuário
        2º) data de pagamento do código (mais antiga -> mais recente)
        3º) ordem da linha na planilha (desempate quando datas
            iguais/ausentes)
 5. O PDF final é comprimido (compactação de imagens/streams) para reduzir
    o tamanho do arquivo.
 6. O arquivo final é nomeado como:
       "<prefixo da nota> ND <número>-<ano atual>.pdf"
    Ex.: "Nota de débito - Neon Consiga+ ND 1266-2026.pdf"
    O número do ND usado no nome é exatamente o valor da coluna ND da
    planilha (não o número extraído do arquivo de nota).
 7. Códigos sem ND definido na base continuam sendo processados
    individualmente (comportamento antigo, com nome vindo da coluna de
    nome/lançamento).
 8. Ao final, é SEMPRE gerado um arquivo de LOG (.txt) na pasta de saída,
    mesmo que ocorram erros durante o processamento - com um resumo do
    que foi feito, avisos e a lista de arquivos possivelmente ignorados.
"""

from pathlib import Path
from datetime import datetime, date
import re
import csv
import io
import unicodedata
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from PIL import Image

try:
    import pikepdf
    PIKEPDF_DISPONIVEL = True
except ImportError:
    PIKEPDF_DISPONIVEL = False

# ----------------------------
# Constantes / configuração
# ----------------------------
PADRAO_ARQUIVO = re.compile(r"^(\d+)([a-z]?)(\d*)\.pdf$", re.IGNORECASE)
PADRAO_NOTA_PREFIXO = re.compile(r"^nota de debito\b")
PREFIXO_NOTA_PADRAO = "Nota de débito"

QUALIDADE_JPEG = 60            # qualidade (1-95) usada ao recomprimir imagens
RESOLUCAO_MAXIMA_PX = 2000     # maior dimensão (px) permitida para imagens

FORMATOS_DATA = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d.%m.%Y"]


# ============================================================
# Seleções (GUI)
# ============================================================
def selecionar_entrada_saida_e_base():
    root = tk.Tk()
    root.withdraw()

    entrada = filedialog.askdirectory(title="Pasta de entrada")
    if not entrada:
        return None, None, None

    saida = filedialog.askdirectory(title="Pasta de saída")
    if not saida:
        return None, None, None

    base = filedialog.askopenfilename(
        title="Base de nomenclatura (Id | LANÇAMENTO | DATA DO PAGAMENTO | ND)",
        filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
    )
    if not base:
        return None, None, None

    return Path(entrada), Path(saida), Path(base)


def perguntar_ordem_popup(letras):
    root = tk.Tk()
    root.withdraw()

    letras_str = ", ".join(sorted(letras))

    ordem = simpledialog.askstring(
        "Ordem dos PDFs",
        f"Letras encontradas: {letras_str}\n\nDigite a ordem (ex: C,G,A):"
    )

    if not ordem:
        return None

    return [x.strip().upper() for x in ordem.split(",") if x.strip()]


# ============================================================
# Utilidades de texto
# ============================================================
def normalizar_nome(nome):
    return re.sub(r"\.pdf$", "", str(nome).strip(), flags=re.IGNORECASE)


def sanitizar(nome):
    nome = re.sub(r'[<>:"/\\|?*]', "_", str(nome))
    return re.sub(r"\s+", " ", nome).strip()


def normalizar_texto(s):
    """minúsculas, sem acento, espaços normalizados - só para comparação."""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def extrair_letra(nome):
    m = re.match(r"^\d+([a-z])", nome.lower())
    return m.group(1).upper() if m else ""


def extrair_numero_extra(nome):
    m = PADRAO_ARQUIVO.match(nome)
    if not m:
        return 0
    numero = m.group(3)
    return int(numero) if numero else 0


def extrair_codigo_do_nome(nome):
    m = PADRAO_ARQUIVO.match(nome)
    return m.group(1) if m else None


def limpar_nd(valor):
    """Normaliza o valor da coluna ND lido da base (número ou texto)."""
    if valor is None:
        return ""
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    if isinstance(valor, int):
        return str(valor)
    s = str(valor).strip()
    if not s:
        return ""
    m = re.match(r"^(\d+)\.0$", s)
    if m:
        return m.group(1)
    return s


def nd_chave(valor):
    """Chave normalizada para comparar ND (somente dígitos, sem zeros à esquerda)."""
    digitos = re.sub(r"\D", "", str(valor or ""))
    if not digitos:
        return ""
    try:
        return str(int(digitos))
    except ValueError:
        return digitos


def parse_data(valor):
    """
    Converte o valor lido da coluna 'DATA DO PAGAMENTO' em um objeto
    datetime.date. Aceita datetime/date já convertidos pelo openpyxl,
    número serial do Excel (quando a célula não está formatada como
    data) e texto em vários formatos comuns (dd/mm/aaaa, aaaa-mm-dd...).
    Retorna None quando não for possível interpretar.
    """
    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor

    if isinstance(valor, (int, float)):
        try:
            convertido = from_excel(valor)
            if isinstance(convertido, datetime):
                return convertido.date()
            if isinstance(convertido, date):
                return convertido
        except Exception:
            pass
        return None

    s = str(valor).strip()
    if not s:
        return None

    for fmt in FORMATOS_DATA:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue

    # data com hora junto (ex: "15/01/2026 00:00:00") -> tenta só a parte da data
    m = re.match(r"^(\d{1,4}[/-]\d{1,2}[/-]\d{1,4})", s)
    if m and m.group(1) != s:
        return parse_data(m.group(1))

    return None


# ============================================================
# Identificação de arquivo de NOTA DE DÉBITO
# ============================================================
def eh_arquivo_nota(nome_arquivo):
    base = re.sub(r"\.pdf$", "", str(nome_arquivo).strip(), flags=re.IGNORECASE)
    return bool(PADRAO_NOTA_PREFIXO.match(normalizar_texto(base)))


def extrair_nd_do_nome_nota(nome_arquivo):
    base = re.sub(r"\.pdf$", "", str(nome_arquivo).strip(), flags=re.IGNORECASE)
    m = re.search(r"\bnd\b\s*[-:\s]*(\d+)", base, flags=re.IGNORECASE)
    return m.group(1) if m else None


def extrair_prefixo_nota(nome_arquivo):
    """
    Retorna o texto do nome do arquivo de nota que vem ANTES de 'ND'
    (preservando a marca/variação, ex: 'Nota de débito - Neon Consiga+').
    Se não conseguir identificar, cai no prefixo padrão.
    """
    base = re.sub(r"\.pdf$", "", str(nome_arquivo).strip(), flags=re.IGNORECASE)
    m = re.search(r"^(.*?)\bnd\b", base, flags=re.IGNORECASE)
    if m:
        prefixo = m.group(1).strip(" -\u2013\u2014")
        if prefixo:
            return prefixo
    return PREFIXO_NOTA_PADRAO


# ============================================================
# Base de nomenclatura (Excel/CSV)
# ============================================================
def detectar_colunas(linha_cabecalho):
    """
    Tenta identificar, pelos nomes das colunas de uma linha de cabeçalho,
    os índices de: id, lancamento, data_pagamento, nd.
    Retorna um dict (pode conter só parte das chaves).
    """
    mapeamento = {}
    for idx, valor in enumerate(linha_cabecalho):
        norm = normalizar_texto(valor or "")
        if norm in ("id", "codigo"):
            mapeamento.setdefault("id", idx)
        elif "lancamento" in norm or "lançamento" in str(valor or "").lower():
            mapeamento.setdefault("lancamento", idx)
        elif "data" in norm and "pagamento" in norm:
            mapeamento.setdefault("data_pagamento", idx)
        elif norm == "nd":
            mapeamento.setdefault("nd", idx)
    return mapeamento


def _ler_texto_csv(caminho):
    """
    Lê o conteúdo de um .csv tentando várias codificações comuns, já que
    o CSV pode ter sido exportado pelo Excel/Windows em português (que
    normalmente NÃO gera UTF-8, e sim 'cp1252'/'latin-1' ou, mais raro,
    'utf-16'). Tenta na ordem: utf-8-sig, utf-8, cp1252, latin-1.
    """
    dados = caminho.read_bytes()

    # UTF-16 (Excel "CSV UTF-16" / "Unicode Text") tem BOM característico.
    if dados[:2] in (b"\xff\xfe", b"\xfe\xff"):
        return dados.decode("utf-16")

    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return dados.decode(enc)
        except UnicodeDecodeError:
            continue

    # último recurso: nunca falha, mas pode substituir caracteres inválidos
    return dados.decode("utf-8", errors="replace")


def _detectar_delimitador_csv(texto):
    """
    Detecta o separador usado no CSV (',', ';' ou tabulação - o Excel em
    português normalmente exporta CSV separado por ';'). Usa csv.Sniffer
    como primeira tentativa e cai para contagem de caracteres na primeira
    linha caso o Sniffer não consiga decidir.
    """
    amostra = "\n".join(texto.splitlines()[:5])
    try:
        return csv.Sniffer().sniff(amostra, delimiters=",;\t").delimiter
    except csv.Error:
        pass

    primeira_linha = texto.splitlines()[0] if texto.splitlines() else ""
    contagens = {d: primeira_linha.count(d) for d in (";", ",", "\t")}
    delimitador = max(contagens, key=contagens.get)
    return delimitador if contagens[delimitador] > 0 else ","


def _linhas_planilha(caminho):
    if caminho.suffix.lower() == ".csv":
        texto = _ler_texto_csv(caminho)
        delimitador = _detectar_delimitador_csv(texto)
        reader = csv.reader(io.StringIO(texto), delimiter=delimitador)
        return [list(row) for row in reader if any(c not in (None, "") for c in row)]

    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    return [
        list(row) for row in ws.iter_rows(values_only=True)
        if row and any(c not in (None, "") for c in row)
    ]


def carregar_base(caminho):
    """
    Retorna um dict: chave_codigo (normalizado, minúsculo) -> {
        "nome": nome de saída (fallback quando não há ND),
        "nd": número do ND (string limpa, pode ser ""),
        "data_pagamento": datetime.date ou None,
        "ordem_planilha": int (posição da linha na planilha, para desempate),
    }
    Suporta tanto a base NOVA (com cabeçalho: Id / LANÇAMENTO /
    DATA DO PAGAMENTO / ND, em qualquer ordem de colunas) quanto a base
    ANTIGA (sem cabeçalho: código / nome / ND).
    """
    linhas = _linhas_planilha(caminho)
    if not linhas:
        return {}

    primeira_linha = linhas[0]
    colunas = detectar_colunas(primeira_linha)
    usa_cabecalho = "id" in colunas and "nd" in colunas

    if usa_cabecalho:
        idx_id = colunas["id"]
        idx_nome = colunas.get("lancamento")
        idx_data = colunas.get("data_pagamento")
        idx_nd = colunas["nd"]
        linhas_dados = linhas[1:]
    else:
        idx_id, idx_nome, idx_data, idx_nd = 0, 1, None, 2
        linhas_dados = linhas

    def obter(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    mapa = {}
    ordem = 0
    for row in linhas_dados:
        codigo_raw = obter(row, idx_id)
        if codigo_raw in (None, ""):
            continue

        chave = normalizar_nome(codigo_raw).lower()
        if not chave:
            continue

        nome_raw = obter(row, idx_nome)
        nome = sanitizar(nome_raw) if nome_raw else ""
        nd_raw = obter(row, idx_nd)
        nd = limpar_nd(nd_raw)
        data_pagamento = parse_data(obter(row, idx_data)) if idx_data is not None else None

        mapa[chave] = {
            "nome": nome or chave,
            "nd": nd,
            "data_pagamento": data_pagamento,
            "ordem_planilha": ordem,
        }
        ordem += 1

    return mapa


# ============================================================
# Ordenação dos comprovantes dentro de um mesmo ND
# ============================================================
def ordenar_itens_do_nd(arquivos, mapa, ordem_usuario):
    """
    Ordena os arquivos (comprovantes) de um mesmo ND por:
      1) posição da letra do arquivo na ordem definida pelo usuário
      2) data de pagamento do código a que o arquivo pertence
         (mais antiga -> mais recente)
      3) ordem da linha na planilha (desempate: datas iguais/ausentes,
         respeita a ordem em que os códigos aparecem na base)
      4) número extra no nome do arquivo (ex: "100a2.pdf")
      5) nome do arquivo (desempate final/estável)
    """
    def chave(arq):
        nome = arq.name
        letra = extrair_letra(nome)
        codigo = extrair_codigo_do_nome(nome) or ""
        info = mapa.get(codigo.lower(), {})

        pos_letra = ordem_usuario.index(letra) if (ordem_usuario and letra in ordem_usuario) else 999

        data_pg = info.get("data_pagamento")
        sem_data = data_pg is None
        ordem_planilha = info.get("ordem_planilha", 10 ** 9)
        num_extra = extrair_numero_extra(nome)

        return (pos_letra, sem_data, data_pg or date.max, ordem_planilha, num_extra, nome.lower())

    return sorted(arquivos, key=chave)


def ordenar_arquivos_individual(arquivos, ordem_usuario):
    """Ordenação usada para códigos SEM ND (fallback individual) - só pela letra/número."""
    def chave(arq):
        nome = arq.name
        letra = extrair_letra(nome)
        pos_letra = ordem_usuario.index(letra) if (ordem_usuario and letra in ordem_usuario) else 999
        num_extra = extrair_numero_extra(nome)
        return (pos_letra, num_extra, nome.lower())

    return sorted(arquivos, key=chave)


# ============================================================
# Nome final (fallback para códigos SEM ND na base)
# ============================================================
def obter_nome_saida(codigo, arquivos, mapa):
    candidatos = [codigo]

    for arq in arquivos:
        candidatos.append(normalizar_nome(arq.name))

    for c in candidatos:
        info = mapa.get(c.lower())
        if info and info.get("nome"):
            return sanitizar(info["nome"])

    return codigo


def montar_nome_nota_unificada(prefixo, nd_original, ano):
    return f"{sanitizar(prefixo)} ND {sanitizar(nd_original)}-{ano}"


# ============================================================
# Evitar sobrescrever
# ============================================================
def nome_unico(path):
    if not path.exists():
        return path

    i = 1
    while True:
        novo = path.with_name(f"{path.stem}_{i}{path.suffix}")
        if not novo.exists():
            return novo
        i += 1


# ============================================================
# Leitura / montagem de páginas
# ============================================================
def adicionar_paginas(writer, caminho):
    reader = PdfReader(str(caminho))
    for pagina in reader.pages:
        writer.add_page(pagina)


# ============================================================
# Compressão do PDF final
# ============================================================
def comprimir_pdf(writer):
    """Reduz o tamanho do PDF: recomprime imagens e streams de conteúdo."""
    for page in writer.pages:
        try:
            for img in list(page.images):
                try:
                    pil = img.image
                    if pil.mode in ("RGBA", "LA", "P"):
                        pil = pil.convert("RGB")
                    elif pil.mode == "1":
                        pil = pil.convert("L")

                    if max(pil.size) > RESOLUCAO_MAXIMA_PX:
                        escala = RESOLUCAO_MAXIMA_PX / max(pil.size)
                        novo_tam = (
                            max(1, int(pil.size[0] * escala)),
                            max(1, int(pil.size[1] * escala)),
                        )
                        pil = pil.resize(novo_tam, Image.LANCZOS)

                    img.replace(pil, quality=QUALIDADE_JPEG)
                except Exception:
                    # se a imagem não puder ser recomprimida, mantém original
                    continue

            page.compress_content_streams()
        except Exception:
            continue

    try:
        writer.compress_identical_objects()
    except Exception:
        pass


def salvar_pdf_comprimido(writer, destino):
    """Escreve o PDF no destino, tentando uma segunda passada de compressão via pikepdf."""
    buffer = io.BytesIO()
    writer.write(buffer)
    dados = buffer.getvalue()

    if PIKEPDF_DISPONIVEL:
        try:
            with pikepdf.open(io.BytesIO(dados)) as pdf:
                pdf.save(
                    str(destino),
                    compress_streams=True,
                    recompress_flate=True,
                    object_stream_mode=pikepdf.ObjectStreamMode.generate,
                )
            return
        except Exception:
            pass  # cai para escrita simples abaixo

    with open(destino, "wb") as f:
        f.write(dados)


# ============================================================
# Escaneamento da pasta de entrada
# ============================================================
def escanear_pasta(pasta_entrada):
    """
    Retorna:
        grupos: dict codigo -> [Path, ...]  (comprovantes)
        notas_arquivo: dict nd_chave -> Path  (arquivo de nota daquele ND)
        letras_detectadas: set de letras encontradas nos comprovantes
        ignorados: list[(nome_arquivo, motivo)]
        total_lidos: int (quantidade de PDFs encontrados na pasta)
    """
    grupos = {}
    notas_arquivo = {}
    letras_detectadas = set()
    ignorados = []
    total_lidos = 0

    for arq in sorted(pasta_entrada.iterdir()):
        if not (arq.is_file() and arq.suffix.lower() == ".pdf"):
            continue

        total_lidos += 1
        nome = arq.name.strip()

        if eh_arquivo_nota(nome):
            nd_arquivo = extrair_nd_do_nome_nota(nome)
            if not nd_arquivo:
                ignorados.append(
                    (nome, "Arquivo de nota de débito sem número de ND identificável no nome")
                )
                continue

            chave = nd_chave(nd_arquivo)
            if chave in notas_arquivo:
                ignorados.append(
                    (nome, f"Nota duplicada para o ND {nd_arquivo} (mantida a primeira encontrada)")
                )
            else:
                notas_arquivo[chave] = arq
            continue

        m = PADRAO_ARQUIVO.match(nome)
        if not m:
            ignorados.append(
                (nome, "Nome de arquivo fora do padrão esperado (não é comprovante nem nota)")
            )
            continue

        codigo = m.group(1)

        letra = extrair_letra(nome)
        if letra:
            letras_detectadas.add(letra)

        grupos.setdefault(codigo, []).append(arq)

    return grupos, notas_arquivo, letras_detectadas, ignorados, total_lidos


# ============================================================
# Montagem dos grupos de ND (a partir da base)
# ============================================================
def montar_grupos_nd(mapa):
    """
    A partir da base carregada, monta um dict:
        nd_chave -> {"nd_original": str, "codigos": [codigo, ...]}
    preservando a ordem de aparição dos códigos na base.
    Apenas códigos com ND preenchido entram aqui.
    """
    grupos_nd = {}
    for codigo_key, info in mapa.items():
        nd_original = (info.get("nd") or "").strip()
        if not nd_original:
            continue

        chave = nd_chave(nd_original)
        if not chave:
            continue

        grupo = grupos_nd.setdefault(chave, {"nd_original": nd_original, "codigos": []})
        if codigo_key not in grupo["codigos"]:
            grupo["codigos"].append(codigo_key)

    return grupos_nd


# ============================================================
# Processamento principal (lógica pura, sem GUI - testável)
# ============================================================
def processar(grupos, notas_arquivo, mapa, pasta_saida, ordem_usuario, ignorados_iniciais):
    """
    Gera as notas de débito unificadas por ND e, para os códigos sem ND
    definido na base, mantém o comportamento antigo (merge individual).

    Retorna: gerados (list[dict]), ignorados (list[tuple]), avisos (list[str])
    """
    ignorados = list(ignorados_iniciais)
    avisos = []
    gerados = []
    codigos_processados = set()
    ano_atual = datetime.now().year

    grupos_nd = montar_grupos_nd(mapa)

    # ------------------------------------------------------------------
    # 1) Uma nota de débito unificada por ND existente na base
    # ------------------------------------------------------------------
    for chave_nd, info_nd in grupos_nd.items():
        nd_original = info_nd["nd_original"]
        codigos_do_nd = info_nd["codigos"]

        writer = PdfWriter()
        incluidos = []
        codigos_sem_arquivo = []
        prefixo_nota = PREFIXO_NOTA_PADRAO

        nota_path = notas_arquivo.get(chave_nd)
        if nota_path:
            try:
                adicionar_paginas(writer, nota_path)
                incluidos.append(nota_path.name)
                prefixo_nota = extrair_prefixo_nota(nota_path.name)
            except Exception as e:
                avisos.append(f"ND {nd_original}: erro ao ler o arquivo de nota '{nota_path.name}': {e}")
        else:
            avisos.append(f"ND {nd_original}: arquivo de nota de débito não encontrado na pasta de entrada.")

        # junta todos os arquivos de todos os códigos deste ND e ordena
        # de acordo com: letra -> data de pagamento -> ordem da planilha
        arquivos_do_nd = []
        for codigo in codigos_do_nd:
            if codigo in codigos_processados:
                avisos.append(
                    f"ND {nd_original}: código '{codigo}' já havia sido atribuído a outro ND na base; "
                    f"ignorado aqui para evitar duplicidade."
                )
                continue

            arquivos_codigo = grupos.get(codigo)
            if not arquivos_codigo:
                codigos_sem_arquivo.append(codigo)
                continue

            arquivos_do_nd.extend(arquivos_codigo)
            codigos_processados.add(codigo)

        try:
            for arq in ordenar_itens_do_nd(arquivos_do_nd, mapa, ordem_usuario):
                adicionar_paginas(writer, arq)
                incluidos.append(arq.name)
        except Exception as e:
            avisos.append(f"ND {nd_original}: erro ao ler arquivo(s) de comprovantes: {e}")

        if codigos_sem_arquivo:
            avisos.append(
                f"ND {nd_original}: código(s) da base sem arquivo correspondente na pasta: "
                + ", ".join(codigos_sem_arquivo)
            )

        if len(writer.pages) == 0:
            avisos.append(
                f"ND {nd_original}: nenhuma página encontrada (nem nota, nem comprovantes) - "
                f"nota unificada NÃO foi gerada."
            )
            continue

        try:
            comprimir_pdf(writer)

            nome_final = montar_nome_nota_unificada(prefixo_nota, nd_original, ano_atual)
            destino = nome_unico(pasta_saida / f"{nome_final}.pdf")
            salvar_pdf_comprimido(writer, destino)

            gerados.append({
                "arquivo": destino.name,
                "nd": nd_original,
                "paginas": len(writer.pages),
                "qtd_arquivos": len(incluidos),
                "incluidos": incluidos,
            })
        except Exception as e:
            avisos.append(f"ND {nd_original}: erro ao gerar/salvar o PDF unificado: {e}")

    # ------------------------------------------------------------------
    # 2) Notas de débito encontradas na pasta cujo ND não existe na base
    # ------------------------------------------------------------------
    for chave_nd, nota_path in notas_arquivo.items():
        if chave_nd not in grupos_nd:
            ignorados.append(
                (nota_path.name, "Nota de débito encontrada, mas o ND não consta na base "
                                 "(nenhum código associado a esse ND).")
            )

    # ------------------------------------------------------------------
    # 3) Códigos sem ND definido na base (fallback - merge individual)
    # ------------------------------------------------------------------
    for codigo, arquivos in grupos.items():
        if codigo in codigos_processados:
            continue

        try:
            arquivos_ordenados = ordenar_arquivos_individual(arquivos, ordem_usuario)

            writer = PdfWriter()
            for arq in arquivos_ordenados:
                adicionar_paginas(writer, arq)

            comprimir_pdf(writer)

            nome_saida = obter_nome_saida(codigo, arquivos_ordenados, mapa)
            destino = nome_unico(pasta_saida / f"{nome_saida}.pdf")
            salvar_pdf_comprimido(writer, destino)

            info_base = mapa.get(codigo.lower())
            motivo = ("código não encontrado na base"
                      if not info_base else "ND não definido na base para este código")
            avisos.append(f"Código '{codigo}': {motivo}; processado individualmente (sem agrupar por ND).")

            gerados.append({
                "arquivo": destino.name,
                "nd": "",
                "paginas": len(writer.pages),
                "qtd_arquivos": len(arquivos_ordenados),
                "incluidos": [a.name for a in arquivos_ordenados],
            })
        except Exception as e:
            avisos.append(f"Código '{codigo}': erro ao gerar PDF individual: {e}")

    return gerados, ignorados, avisos


# ============================================================
# Log
# ============================================================
def gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos, erro_fatal=None):
    agora = datetime.now()
    caminho_log = pasta_saida / f"LOG_processamento_{agora.strftime('%Y%m%d_%H%M%S')}.txt"

    linhas = []
    linhas.append("=" * 72)
    linhas.append("LOG DE PROCESSAMENTO - UNIFICAÇÃO DE PDFs / NOTAS DE DÉBITO")
    linhas.append(f"Data/Hora: {agora.strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append("=" * 72)
    linhas.append("")

    if erro_fatal:
        linhas.append("ERRO FATAL - O PROCESSAMENTO FOI INTERROMPIDO")
        linhas.append("-" * 72)
        linhas.append(erro_fatal)
        linhas.append("")

    linhas.append("RESUMO GERAL")
    linhas.append("-" * 72)
    linhas.append(f"Arquivos PDF lidos na pasta de entrada .......... {total_lidos}")
    linhas.append(f"PDFs unificados gerados (por ND ou individual) .. {len(gerados)}")
    linhas.append(f"Arquivos possivelmente ignorados ................ {len(ignorados)}")
    linhas.append(f"Avisos gerados durante o processamento .......... {len(avisos)}")
    linhas.append("")

    linhas.append("ARQUIVOS GERADOS")
    linhas.append("-" * 72)
    if gerados:
        for g in gerados:
            rotulo_nd = f"ND {g['nd']}" if g["nd"] else "sem ND (individual)"
            linhas.append(
                f"- {g['arquivo']}  [{rotulo_nd} | {g['paginas']} página(s) | "
                f"{g['qtd_arquivos']} arquivo(s) unidos]"
            )
            for nome_incluido in g["incluidos"]:
                linhas.append(f"    • {nome_incluido}")
    else:
        linhas.append("(nenhum arquivo foi gerado)")
    linhas.append("")

    linhas.append("AVISOS")
    linhas.append("-" * 72)
    if avisos:
        for a in avisos:
            linhas.append(f"- {a}")
    else:
        linhas.append("(nenhum aviso)")
    linhas.append("")

    linhas.append("ARQUIVOS POSSIVELMENTE IGNORADOS")
    linhas.append("-" * 72)
    if ignorados:
        for nome, motivo in ignorados:
            linhas.append(f"- {nome}")
            linhas.append(f"    Motivo: {motivo}")
    else:
        linhas.append("(nenhum arquivo foi ignorado)")
    linhas.append("")

    linhas.append("=" * 72)
    linhas.append("Fim do log.")

    caminho_log.write_text("\n".join(linhas), encoding="utf-8")
    return caminho_log


# ============================================================
# MAIN (GUI)
# ============================================================
def main():
    pasta_entrada, pasta_saida, base = selecionar_entrada_saida_e_base()
    if not pasta_entrada:
        return

    pasta_saida.mkdir(exist_ok=True)

    # A partir deste ponto, SEMPRE geramos um log, mesmo se algo falhar,
    # para que o usuário consiga diagnosticar o que aconteceu.
    gerados, ignorados, avisos = [], [], []
    total_lidos = 0
    erro_fatal = None

    try:
        mapa = carregar_base(base)

        grupos, notas_arquivo, letras_detectadas, ignorados, total_lidos = escanear_pasta(pasta_entrada)

        if not grupos and not notas_arquivo:
            messagebox.showinfo("Resultado", "Nenhum PDF válido encontrado (nem comprovantes, nem notas).")
            gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos,
                       erro_fatal="Nenhum PDF válido encontrado na pasta de entrada.")
            return

        ordem_usuario = []
        if letras_detectadas:
            ordem_usuario = perguntar_ordem_popup(letras_detectadas)
            if not ordem_usuario:
                messagebox.showwarning("Aviso", "Ordem não definida. Operação cancelada.")
                gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos,
                           erro_fatal="Operação cancelada pelo usuário (ordem das letras não informada).")
                return

        gerados, ignorados, avisos = processar(
            grupos, notas_arquivo, mapa, pasta_saida, ordem_usuario, ignorados
        )

    except Exception:
        erro_fatal = traceback.format_exc()

    caminho_log = gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos, erro_fatal=erro_fatal)

    if erro_fatal:
        messagebox.showerror(
            "Erro durante o processamento",
            f"Ocorreu um erro e o processamento foi interrompido.\n\n"
            f"Detalhes salvos no log:\n{caminho_log}"
        )
    else:
        messagebox.showinfo(
            "Concluído",
            f"{len(gerados)} PDF(s) gerado(s).\n"
            f"{len(ignorados)} arquivo(s) possivelmente ignorado(s).\n"
            f"{len(avisos)} aviso(s).\n\n"
            f"Log salvo em:\n{caminho_log}"
        )

    print(f"\nLog salvo em: {caminho_log}")
    print("\nGerados:")
    for g in gerados:
        print(" -", g["arquivo"])

    if ignorados:
        print("\nIgnorados:")
        for nome, motivo in ignorados:
            print(" -", nome, "->", motivo)

    if avisos:
        print("\nAvisos:")
        for a in avisos:
            print(" -", a)

    if erro_fatal:
        print("\nERRO FATAL:")
        print(erro_fatal)


# ----------------------------
if __name__ == "__main__":
    main()
