"""
Unificador de PDFs / Notas de Débito - v4.0
=============================================

Fluxo:
 1. Lê os comprovantes PDF de uma pasta de entrada (padrão de nome:
    "<código><letra opcional><número opcional>.pdf", ex: "100.pdf", "100a.pdf").
 2. Lê também os arquivos de NOTA DE DÉBITO, cujo nome começa com
    "Nota de débito - Neon Pagamentos ND" (ex: "Nota de débito - Neon
    Pagamentos ND 1262.pdf"). O número do ND é extraído do próprio nome
    do arquivo.
 3. Lê uma base (Excel/CSV) com 3 colunas:
       Coluna A -> código do arquivo (ex: 100)
       Coluna B -> nome de saída (usado apenas quando o código NÃO tiver
                   ND definido na base - modo de compatibilidade)
       Coluna C -> ND (número da nota de débito à qual o código pertence)
 4. Para cada número de ND presente na base, gera UM ÚNICO PDF unificado
    contendo: o arquivo da nota de débito daquele ND + todos os
    comprovantes de todos os códigos vinculados àquele ND (na ordem em
    que aparecem na base, e dentro de cada código na ordem de letras
    definida pelo usuário).
 5. O PDF final é comprimido (compactação de imagens/streams) para reduzir
    o tamanho do arquivo.
 6. O arquivo final é nomeado como:
       "Nota de débito - Neon Pagamentos ND <número>-<ano atual>.pdf"
 7. Códigos sem ND definido na base continuam sendo processados
    individualmente (comportamento antigo, com nome vindo da coluna B).
 8. Ao final, é gerado um arquivo de LOG (.txt) na pasta de saída, com um
    resumo do que foi feito e a lista de arquivos possivelmente ignorados.
"""

from pathlib import Path
from datetime import datetime
import re
import csv
import io
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

from pypdf import PdfReader, PdfWriter
from openpyxl import load_workbook
from PIL import Image

try:
    import pikepdf
    PIKEPDF_DISPONIVEL = True
except ImportError:
    PIKEPDF_DISPONIVEL = False

# ----------------------------
# Constantes / configuração de compressão
# ----------------------------
PADRAO_ARQUIVO = re.compile(r"^(\d+)([a-z]?)(\d*)\.pdf$", re.IGNORECASE)
PADRAO_NOTA_PREFIXO = re.compile(r"^nota de debito\s*-\s*neon pagamentos\s+nd")

QUALIDADE_JPEG = 60           # qualidade (1-95) usada ao recomprimir imagens
RESOLUCAO_MAXIMA_PX = 2000     # maior dimensão (px) permitida para imagens

# ============================================================
# Seleções (GUI)
# ============================================================
def selecionar_entrada_saida_e_base():
    root = tk.Tk()
    root.withdraw()

    entrada = filedialog.askdirectory(title="Pasta de entrada")
    if not entrada:
        return None, None, None

    saida = filedialog.askdirectory(title="Pasta de saída")
    if not saida:
        return None, None, None

    base = filedialog.askopenfilename(
        title="Base de nomenclatura (Coluna A: código | B: nome | C: ND)",
        filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")]
    )
    if not base:
        return None, None, None

    return Path(entrada), Path(saida), Path(base)


def perguntar_ordem_popup(letras):
    root = tk.Tk()
    root.withdraw()

    letras_str = ", ".join(sorted(letras))

    ordem = simpledialog.askstring(
        "Ordem dos PDFs",
        f"Letras encontradas: {letras_str}\n\nDigite a ordem (ex: C,G,A):"
    )

    if not ordem:
        return None

    return [x.strip().upper() for x in ordem.split(",") if x.strip()]


# ============================================================
# Utilidades de texto
# ============================================================
def normalizar_nome(nome):
    return re.sub(r"\.pdf$", "", str(nome).strip(), flags=re.IGNORECASE)


def sanitizar(nome):
    nome = re.sub(r'[<>:"/\\|?*]', "_", str(nome))
    return re.sub(r"\s+", " ", nome).strip()


def normalizar_texto(s):
    """minúsculas, sem acento, espaços normalizados - só para comparação."""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def extrair_letra(nome):
    m = re.match(r"^\d+([a-z])", nome.lower())
    return m.group(1).upper() if m else ""


def limpar_nd(valor):
    """Normaliza o valor da coluna ND lido da base (número ou texto)."""
    if valor is None:
        return ""
    if isinstance(valor, float):
        if valor.is_integer():
            return str(int(valor))
        return str(valor).strip()
    if isinstance(valor, int):
        return str(valor)
    s = str(valor).strip()
    if not s:
        return ""
    m = re.match(r"^(\d+)\.0$", s)
    if m:
        return m.group(1)
    return s


def nd_chave(valor):
    """Chave normalizada para comparar ND (somente dígitos, sem zeros à esquerda)."""
    digitos = re.sub(r"\D", "", str(valor or ""))
    if not digitos:
        return ""
    try:
        return str(int(digitos))
    except ValueError:
        return digitos


# ============================================================
# Identificação de arquivo de NOTA DE DÉBITO
# ============================================================
def eh_arquivo_nota(nome_arquivo):
    base = re.sub(r"\.pdf$", "", str(nome_arquivo).strip(), flags=re.IGNORECASE)
    return bool(PADRAO_NOTA_PREFIXO.match(normalizar_texto(base)))


def extrair_nd_do_nome_nota(nome_arquivo):
    base = re.sub(r"\.pdf$", "", str(nome_arquivo).strip(), flags=re.IGNORECASE)
    m = re.search(r"nd\s*[-:\s]*(\d+)", base, flags=re.IGNORECASE)
    return m.group(1) if m else None


# ============================================================
# Base de nomenclatura (Excel/CSV)
# ============================================================
def carregar_base(caminho):
    """
    Retorna um dict: chave_codigo (normalizado, minúsculo) -> {
        "nome": nome de saída (fallback quando não há ND),
        "nd": número do ND (string limpa, pode ser ""),
    }
    A ordem de inserção do dict reflete a ordem das linhas na base,
    o que é usado para ordenar os códigos dentro de cada ND.
    """
    mapa = {}

    def registrar(codigo_raw, nome_raw, nd_raw):
        chave = normalizar_nome(codigo_raw).lower()
        if not chave:
            return
        nome = sanitizar(nome_raw) if nome_raw else ""
        nd = limpar_nd(nd_raw)
        mapa[chave] = {"nome": nome or chave, "nd": nd}

    if caminho.suffix.lower() == ".csv":
        with open(caminho, encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for linha in reader:
                if len(linha) >= 2 and str(linha[0]).strip():
                    nd_raw = linha[2] if len(linha) >= 3 else ""
                    registrar(linha[0], linha[1], nd_raw)
    else:
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if row and len(row) >= 2 and row[0] not in (None, ""):
                nd_raw = row[2] if len(row) >= 3 else ""
                registrar(row[0], row[1], nd_raw)

    return mapa


# ============================================================
# Ordem dinâmica dos comprovantes dentro de um mesmo código
# ============================================================
def definir_ordem(nome, ordem_usuario):
    nome = nome.lower()

    m = re.match(r"^(\d+)([a-z]?)(\d*)\.pdf$", nome)
    if not m:
        return 9999

    letra = m.group(2).upper()
    numero = m.group(3)

    if ordem_usuario and letra in ordem_usuario:
        pos = ordem_usuario.index(letra)
    else:
        pos = 999

    num_extra = int(numero) if numero else 0

    return pos * 100 + num_extra


def ordenar_arquivos(arquivos, ordem_usuario):
    return sorted(arquivos, key=lambda x: definir_ordem(x.name, ordem_usuario))


# ============================================================
# Nome final (fallback para códigos SEM ND na base)
# ============================================================
def obter_nome_saida(codigo, arquivos, mapa):
    candidatos = [codigo]

    for arq in arquivos:
        candidatos.append(normalizar_nome(arq.name))

    for c in candidatos:
        info = mapa.get(c.lower())
        if info and info.get("nome"):
            return sanitizar(info["nome"])

    return codigo


def montar_nome_nota_unificada(nd_original, ano):
    return f"Nota de débito - Neon Pagamentos ND {sanitizar(nd_original)}-{ano}"


# ============================================================
# Evitar sobrescrever
# ============================================================
def nome_unico(path):
    if not path.exists():
        return path

    i = 1
    while True:
        novo = path.with_name(f"{path.stem}_{i}{path.suffix}")
        if not novo.exists():
            return novo
        i += 1


# ============================================================
# Leitura / montagem de páginas
# ============================================================
def adicionar_paginas(writer, caminho):
    reader = PdfReader(str(caminho))
    for pagina in reader.pages:
        writer.add_page(pagina)


# ============================================================
# Compressão do PDF final
# ============================================================
def comprimir_pdf(writer):
    """Reduz o tamanho do PDF: recomprime imagens e streams de conteúdo."""
    for page in writer.pages:
        try:
            for img in list(page.images):
                try:
                    pil = img.image
                    if pil.mode in ("RGBA", "LA", "P"):
                        pil = pil.convert("RGB")
                    elif pil.mode == "1":
                        pil = pil.convert("L")

                    if max(pil.size) > RESOLUCAO_MAXIMA_PX:
                        escala = RESOLUCAO_MAXIMA_PX / max(pil.size)
                        novo_tam = (
                            max(1, int(pil.size[0] * escala)),
                            max(1, int(pil.size[1] * escala)),
                        )
                        pil = pil.resize(novo_tam, Image.LANCZOS)

                    img.replace(pil, quality=QUALIDADE_JPEG)
                except Exception:
                    # se a imagem não puder ser recomprimida, mantém original
                    continue

            page.compress_content_streams()
        except Exception:
            continue

    try:
        writer.compress_identical_objects()
    except Exception:
        pass


def salvar_pdf_comprimido(writer, destino):
    """Escreve o PDF no destino, tentando uma segunda passada de compressão via pikepdf."""
    buffer = io.BytesIO()
    writer.write(buffer)
    dados = buffer.getvalue()

    if PIKEPDF_DISPONIVEL:
        try:
            with pikepdf.open(io.BytesIO(dados)) as pdf:
                pdf.save(
                    str(destino),
                    compress_streams=True,
                    recompress_flate=True,
                    object_stream_mode=pikepdf.ObjectStreamMode.generate,
                )
            return
        except Exception:
            pass  # cai para escrita simples abaixo

    with open(destino, "wb") as f:
        f.write(dados)


# ============================================================
# Escaneamento da pasta de entrada
# ============================================================
def escanear_pasta(pasta_entrada):
    """
    Retorna:
        grupos: dict codigo -> [Path, ...]  (comprovantes)
        notas_arquivo: dict nd_chave -> Path  (arquivo de nota daquele ND)
        letras_detectadas: set de letras encontradas nos comprovantes
        ignorados: list[(nome_arquivo, motivo)]
        total_lidos: int (quantidade de PDFs encontrados na pasta)
    """
    grupos = {}
    notas_arquivo = {}
    letras_detectadas = set()
    ignorados = []
    total_lidos = 0

    for arq in sorted(pasta_entrada.iterdir()):
        if not (arq.is_file() and arq.suffix.lower() == ".pdf"):
            continue

        total_lidos += 1
        nome = arq.name.strip()

        if eh_arquivo_nota(nome):
            nd_arquivo = extrair_nd_do_nome_nota(nome)
            if not nd_arquivo:
                ignorados.append(
                    (nome, "Arquivo de nota de débito sem número de ND identificável no nome")
                )
                continue

            chave = nd_chave(nd_arquivo)
            if chave in notas_arquivo:
                ignorados.append(
                    (nome, f"Nota duplicada para o ND {nd_arquivo} (mantida a primeira encontrada)")
                )
            else:
                notas_arquivo[chave] = arq
            continue

        m = PADRAO_ARQUIVO.match(nome)
        if not m:
            ignorados.append(
                (nome, "Nome de arquivo fora do padrão esperado (não é comprovante nem nota)")
            )
            continue

        codigo = m.group(1)

        letra = extrair_letra(nome)
        if letra:
            letras_detectadas.add(letra)

        grupos.setdefault(codigo, []).append(arq)

    return grupos, notas_arquivo, letras_detectadas, ignorados, total_lidos


# ============================================================
# Montagem dos grupos de ND (a partir da base)
# ============================================================
def montar_grupos_nd(mapa):
    """
    A partir da base carregada, monta um dict:
        nd_chave -> {"nd_original": str, "codigos": [codigo, ...]}
    preservando a ordem de aparição dos códigos na base.
    Apenas códigos com ND preenchido entram aqui.
    """
    grupos_nd = {}
    for codigo_key, info in mapa.items():
        nd_original = (info.get("nd") or "").strip()
        if not nd_original:
            continue

        chave = nd_chave(nd_original)
        if not chave:
            continue

        grupo = grupos_nd.setdefault(chave, {"nd_original": nd_original, "codigos": []})
        if codigo_key not in grupo["codigos"]:
            grupo["codigos"].append(codigo_key)

    return grupos_nd


# ============================================================
# Processamento principal (lógica pura, sem GUI - testável)
# ============================================================
def processar(grupos, notas_arquivo, mapa, pasta_saida, ordem_usuario, ignorados_iniciais):
    """
    Gera as notas de débito unificadas por ND e, para os códigos sem ND
    definido na base, mantém o comportamento antigo (merge individual).

    Retorna: gerados (list[dict]), ignorados (list[tuple]), avisos (list[str])
    """
    ignorados = list(ignorados_iniciais)
    avisos = []
    gerados = []
    codigos_processados = set()
    ano_atual = datetime.now().year

    grupos_nd = montar_grupos_nd(mapa)

    # ------------------------------------------------------------------
    # 1) Uma nota de débito unificada por ND existente na base
    # ------------------------------------------------------------------
    for chave_nd, info_nd in grupos_nd.items():
        nd_original = info_nd["nd_original"]
        codigos_do_nd = info_nd["codigos"]

        writer = PdfWriter()
        incluidos = []
        codigos_sem_arquivo = []

        nota_path = notas_arquivo.get(chave_nd)
        if nota_path:
            try:
                adicionar_paginas(writer, nota_path)
                incluidos.append(nota_path.name)
            except Exception as e:
                avisos.append(f"ND {nd_original}: erro ao ler o arquivo de nota '{nota_path.name}': {e}")
        else:
            avisos.append(f"ND {nd_original}: arquivo de nota de débito não encontrado na pasta de entrada.")

        for codigo in codigos_do_nd:
            if codigo in codigos_processados:
                avisos.append(
                    f"ND {nd_original}: código '{codigo}' já havia sido atribuído a outro ND na base; "
                    f"ignorado aqui para evitar duplicidade."
                )
                continue

            arquivos_codigo = grupos.get(codigo)
            if not arquivos_codigo:
                codigos_sem_arquivo.append(codigo)
                continue

            try:
                for arq in ordenar_arquivos(arquivos_codigo, ordem_usuario):
                    adicionar_paginas(writer, arq)
                    incluidos.append(arq.name)
                codigos_processados.add(codigo)
            except Exception as e:
                avisos.append(f"ND {nd_original}: erro ao ler arquivo(s) do código '{codigo}': {e}")

        if codigos_sem_arquivo:
            avisos.append(
                f"ND {nd_original}: código(s) da base sem arquivo correspondente na pasta: "
                + ", ".join(codigos_sem_arquivo)
            )

        if len(writer.pages) == 0:
            avisos.append(
                f"ND {nd_original}: nenhuma página encontrada (nem nota, nem comprovantes) - "
                f"nota unificada NÃO foi gerada."
            )
            continue

        try:
            comprimir_pdf(writer)

            nome_final = montar_nome_nota_unificada(nd_original, ano_atual)
            destino = nome_unico(pasta_saida / f"{nome_final}.pdf")
            salvar_pdf_comprimido(writer, destino)

            gerados.append({
                "arquivo": destino.name,
                "nd": nd_original,
                "paginas": len(writer.pages),
                "qtd_arquivos": len(incluidos),
                "incluidos": incluidos,
            })
        except Exception as e:
            avisos.append(f"ND {nd_original}: erro ao gerar/salvar o PDF unificado: {e}")

    # ------------------------------------------------------------------
    # 2) Notas de débito encontradas na pasta cujo ND não existe na base
    # ------------------------------------------------------------------
    for chave_nd, nota_path in notas_arquivo.items():
        if chave_nd not in grupos_nd:
            ignorados.append(
                (nota_path.name, "Nota de débito encontrada, mas o ND não consta na base "
                                 "(nenhum código associado a esse ND).")
            )

    # ------------------------------------------------------------------
    # 3) Códigos sem ND definido na base (fallback - merge individual)
    # ------------------------------------------------------------------
    for codigo, arquivos in grupos.items():
        if codigo in codigos_processados:
            continue

        try:
            arquivos_ordenados = ordenar_arquivos(arquivos, ordem_usuario)

            writer = PdfWriter()
            for arq in arquivos_ordenados:
                adicionar_paginas(writer, arq)

            comprimir_pdf(writer)

            nome_saida = obter_nome_saida(codigo, arquivos_ordenados, mapa)
            destino = nome_unico(pasta_saida / f"{nome_saida}.pdf")
            salvar_pdf_comprimido(writer, destino)

            info_base = mapa.get(codigo.lower())
            motivo = ("código não encontrado na base"
                      if not info_base else "ND não definido na base para este código")
            avisos.append(f"Código '{codigo}': {motivo}; processado individualmente (sem agrupar por ND).")

            gerados.append({
                "arquivo": destino.name,
                "nd": "",
                "paginas": len(writer.pages),
                "qtd_arquivos": len(arquivos_ordenados),
                "incluidos": [a.name for a in arquivos_ordenados],
            })
        except Exception as e:
            avisos.append(f"Código '{codigo}': erro ao gerar PDF individual: {e}")

    return gerados, ignorados, avisos


# ============================================================
# Log
# ============================================================
def gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos):
    agora = datetime.now()
    caminho_log = pasta_saida / f"LOG_processamento_{agora.strftime('%Y%m%d_%H%M%S')}.txt"

    linhas = []
    linhas.append("=" * 72)
    linhas.append("LOG DE PROCESSAMENTO - UNIFICAÇÃO DE PDFs / NOTAS DE DÉBITO")
    linhas.append(f"Data/Hora: {agora.strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append("=" * 72)
    linhas.append("")

    linhas.append("RESUMO GERAL")
    linhas.append("-" * 72)
    linhas.append(f"Arquivos PDF lidos na pasta de entrada .......... {total_lidos}")
    linhas.append(f"PDFs unificados gerados (por ND ou individual) .. {len(gerados)}")
    linhas.append(f"Arquivos possivelmente ignorados ................ {len(ignorados)}")
    linhas.append(f"Avisos gerados durante o processamento .......... {len(avisos)}")
    linhas.append("")

    linhas.append("ARQUIVOS GERADOS")
    linhas.append("-" * 72)
    if gerados:
        for g in gerados:
            rotulo_nd = f"ND {g['nd']}" if g["nd"] else "sem ND (individual)"
            linhas.append(
                f"- {g['arquivo']}  [{rotulo_nd} | {g['paginas']} página(s) | "
                f"{g['qtd_arquivos']} arquivo(s) unidos]"
            )
            for nome_incluido in g["incluidos"]:
                linhas.append(f"    • {nome_incluido}")
    else:
        linhas.append("(nenhum arquivo foi gerado)")
    linhas.append("")

    linhas.append("AVISOS")
    linhas.append("-" * 72)
    if avisos:
        for a in avisos:
            linhas.append(f"- {a}")
    else:
        linhas.append("(nenhum aviso)")
    linhas.append("")

    linhas.append("ARQUIVOS POSSIVELMENTE IGNORADOS")
    linhas.append("-" * 72)
    if ignorados:
        for nome, motivo in ignorados:
            linhas.append(f"- {nome}")
            linhas.append(f"    Motivo: {motivo}")
    else:
        linhas.append("(nenhum arquivo foi ignorado)")
    linhas.append("")

    linhas.append("=" * 72)
    linhas.append("Fim do log.")

    caminho_log.write_text("\n".join(linhas), encoding="utf-8")
    return caminho_log


# ============================================================
# MAIN (GUI)
# ============================================================
def main():
    pasta_entrada, pasta_saida, base = selecionar_entrada_saida_e_base()
    if not pasta_entrada:
        return

    pasta_saida.mkdir(exist_ok=True)

    try:
        mapa = carregar_base(base)
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível ler a base:\n{e}")
        return

    grupos, notas_arquivo, letras_detectadas, ignorados, total_lidos = escanear_pasta(pasta_entrada)

    if not grupos and not notas_arquivo:
        messagebox.showinfo("Resultado", "Nenhum PDF válido encontrado (nem comprovantes, nem notas).")
        return

    ordem_usuario = []
    if letras_detectadas:
        ordem_usuario = perguntar_ordem_popup(letras_detectadas)
        if not ordem_usuario:
            messagebox.showwarning("Aviso", "Ordem não definida. Operação cancelada.")
            return

    gerados, ignorados, avisos = processar(
        grupos, notas_arquivo, mapa, pasta_saida, ordem_usuario, ignorados
    )

    caminho_log = gerar_log(pasta_saida, gerados, ignorados, avisos, total_lidos)

    messagebox.showinfo(
        "Concluído",
        f"{len(gerados)} PDF(s) gerado(s).\n"
        f"{len(ignorados)} arquivo(s) possivelmente ignorado(s).\n"
        f"{len(avisos)} aviso(s).\n\n"
        f"Log salvo em:\n{caminho_log}"
    )

    print(f"\nLog salvo em: {caminho_log}")
    print("\nGerados:")
    for g in gerados:
        print(" -", g["arquivo"])

    if ignorados:
        print("\nIgnorados:")
        for nome, motivo in ignorados:
            print(" -", nome, "->", motivo)

    if avisos:
        print("\nAvisos:")
        for a in avisos:
            print(" -", a)


# ----------------------------
if __name__ == "__main__":
    main()
